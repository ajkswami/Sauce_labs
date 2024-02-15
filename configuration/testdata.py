import openpyxl

def get_max_row_count(file_path, sheet_name):
    """Get the maximum row count in an Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_row = sheet.max_row
    workbook.close()
    return max_row

def get_max_column_count(file_path, sheet_name):
    """Get the maximum column count in an Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    max_column = sheet.max_column
    workbook.close()
    return max_column

def read_data(file_path, sheet_name, row, column):
    """Read data from a specific cell in an Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    cell_value = sheet.cell(row=row, column=column).value
    workbook.close()
    return cell_value

def write_data(file_path, sheet_name, row, column, value):
    """Write data to a specific cell in an Excel sheet."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(row=row, column=column, value=value)
    workbook.save(file_path)
    workbook.close()

# Example usage:
# file_path = "Test Data.xlsx"
# sheet_name = "Sheet1"
#
# max_row_count = get_max_row_count(file_path, sheet_name)
# max_column_count = get_max_column_count(file_path, sheet_name)
#
# print(f"Max Row Count: {max_row_count}")
# print(f"Max Column Count: {max_column_count}")
#
# row = 1
# column = 1
# cell_value = read_data(file_path, sheet_name, row, column)
# print(f"Data at Row {row}, Column {column}: {cell_value}")
#
# write_data(file_path, sheet_name, 1, 2, "New Value")
